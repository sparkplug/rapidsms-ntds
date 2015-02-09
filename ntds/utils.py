from rapidsms.contrib.locations.models import Location
import re
import difflib
from ntds.models import Reporter
from healthmodels.models import HealthProvider
from django.http import HttpResponse
from openpyxl.workbook import Workbook
import types
import openpyxl
import json
from django.utils.safestring import mark_safe
import datetime
from django.core.servers.basehttp import FileWrapper

from django.db.models import Q
import operator
import phonenumbers

def handle_progress(prog,message=None):
    if prog.step:
        loc=parser[str(prog.step.order)](message)
        if loc:
            prog.step.status="C"
            prog.step.giveup_offset=0
            prog.step.save()
            prog.save()




def make_reg(text):

    template = r"(\b%s\b)"
    w_regex = r""
    words=text.split()
    for word in words:
        if len(w_regex):
            if len(word):
                w_regex = w_regex + r"|" + template % re.escape(word.strip())
        else:
            if len(word):
                w_regex += template % re.escape(word.strip())
    return w_regex

def parse_location(text,type):
    regex=make_reg(text)
    loc=Location.objects.filter(type__name=type,name__iregex=regex)
    if not loc.exists():
        name_types=map(lambda x:x.lower(),list(Location.objects.filter(type__name=type).values_list("name",flat=True)))
        matches = difflib.get_close_matches(text.lower(), name_types)
        if matches:
            loc=Location.objects.filter(type=type,name__iexact=matches[0])
            return loc[0]
        else:
            return None
    else:
        return loc[0]



def parse_district(message):
    reporter=Reporter.objects.get(healthprovider_ptr=message.connection.contact)


    loc=parse_location(message.text,"district")
    if loc:
        reporter.district=loc
        reporter.healthprovider.active=True
        reporter.save()
    return loc


def parse_subcounty(message):
    reporter=Reporter.objects.get(healthprovider_ptr=message.connection.contact)


    loc=parse_location(message.text,"sub_county")
    if loc:
        reporter.subcounty=loc
        reporter.save()
    return loc




def parse_parish(message):
    healthprovider=HealthProvider.objects.create(active=False)
    connection=message.connection
    connection.contact=healthprovider
    connection.save()
    reporter=Reporter.objects.create(healthprovider_ptr=healthprovider)

    loc=parse_location(message.text,"parish")
    if loc:
        reporter.subcounty=loc
        reporter.save()
    return loc




def default(message=None):
    return None

parser={
    "0":default,
    "1":parse_parish,
    "2":parse_subcounty,
    "3":parse_parish

}


def normalize_value(value):

    if isinstance(value, tuple(openpyxl.shared.NUMERIC_TYPES)):
        return value
    elif isinstance(value, (bool, datetime.date)):
        return value
    elif isinstance(value,types.NoneType):
        return ""
    elif isinstance(value,types.StringType):
        #print "str"+value
        return value
    elif isinstance(value,types.ListType):
        return ", ".join(value)

    elif isinstance(value, unicode):
        #print "unicode"
        #print unicodedata.normalize('NFKD', unicode(value)).encode('ascii', 'ignore')
        #openpyxl  hates unicode asciify
        return repr(value)[2:-1]

    else:
        print value
        return repr(value)


def create_workbook(data,filename,headers):

    wb = Workbook(optimized_write = True)
    ws = wb.create_sheet()
    if headers:
        ws.append(headers)

    for rowx, row in enumerate(data):
        ws.append(map(normalize_value,list(row)))

        #import pdb;pdb.set_trace()



        #for colx, value in enumerate(row):
        #   column_letter = get_column_letter((colx + 1))
        #  ws.cell('%s%s'%(column_letter, (rowx+ 1))).value = value
    #ws.auto_filter = ws.calculate_dimension()
    wb.save(filename)
    return True


class ExcelResponse(HttpResponse):
    """
    This class contains utilities that are used to produce Excel reports from datasets stored in a database or scraped
    from a form.
    """

    def __init__(self, data, output_name='excel_report', headers=None,header=None, write_to_file=False, force_csv=False):
        # Make sure we've got the right type of data to work with
        valid_data = False
        if hasattr(data, '__getitem__'):
            if isinstance(data[0], dict):
                if headers is None:
                    headers = data[0].keys()
                data = [[row[col] for col in headers] for row in data]
                data.insert(0, headers)
            if hasattr(data[0], '__getitem__'):
                valid_data = True
        import StringIO

        output = StringIO.StringIO()
        mimetype = 'application/vnd.ms-excel'



        book_created = create_workbook(data,output_name,headers,)


        #book.save(output_name)
        #output.seek(0)
        if not write_to_file:
            super(ExcelResponse, self).__init__(FileWrapper(open(output_name)),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self['Content-Disposition'] = 'attachment;filename="%s.%s"' % \
                                          (output_name.replace('"', '\"'), "xlsx")

class JSONResponse(HttpResponse):
    status = 200

    def __init__(self, data, status=None,serialized=False):
        if not serialized:
            data_json = json.dumps(data)
        else:
            data_json=data

        super(JSONResponse, self).__init__(data_json,
                                           mimetype='application/json',
                                           status=status or self.status)



def get_all_children(nodes, include_self=False):
    if not nodes:
        return Location.tree.none()
    filters = []
    for n in nodes:
        lft, rght = n.lft, n.rght
        if include_self:
            lft -=1
            rght += 1
        filters.append(Q(tree_id=n.tree_id, lft__gt=lft, rght__lt=rght))
    q = reduce(operator.or_, filters)
    return Location.tree.filter(q)

def validate_number(mobile,code="UG"):
    try:
        number=phonenumbers.parse(mobile, code)
    except phonenumbers.NumberParseException:
        return (False,mobile)
    number_str=str(number.country_code)+str(number.national_number)
    if  phonenumbers.is_valid_number(number):
        return (True,number_str)
    else:
        return (False,number_str)


def empty(value):
    if bool(str(value).strip()):
        return False
    else:
        return True

def parse_mobile(value):
    is_valid,mobile=validate_number(value)
    if is_valid:
        return mobile
    else:
        return None







